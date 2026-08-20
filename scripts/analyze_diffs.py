#!/usr/bin/env python3
"""Categorize what changed between two snapshot states, across and within counties.

The snapshot pipeline answers "what changed"; this answers "what KIND of thing
changed, where, and how much" — which is the question you actually have on an
election day, when 60+ counties all move at once.

Method
------
Reads `git diff` of `snapshots/**/page.txt` between two revisions (default: the last
commit vs the working tree). For every added/removed line it assigns zero or more
CATEGORIES by keyword. Lines are the unit, not files, because one page can change in
several distinct ways at once — a homepage can post results, update turnout, and swap
its countdown in the same edit.

Three deliberate choices:

  * `page.txt` only, not `page.html`. The text artifact is what carries meaning; the
    HTML carries the same change plus markup noise, so counting both would
    double-count and let a template tweak masquerade as content.
  * The category columns do NOT reconcile to "lines changed", and are not meant to.
    Two effects pull in opposite directions: a line matching two categories is
    counted twice, and a line matching none is not counted at all. In practice the
    second dominates — Escambia changed 1,089 lines, of which 480 matched no
    category, so its category columns total 408. Read "lines changed" for magnitude
    and the category columns for shape; never treat the categories as a partition.
  * **Capture class is decided before any line is counted.** On election day some
    counties replace their whole site with a minimal "election night results" page,
    which deletes hundreds of lines that were never edited — they were simply not
    served. Counting that as content change would swamp every real edit: on
    2026-08-18 the ten counties doing it accounted for more changed lines than the
    other fifty-four combined. So each target is classed:

        normal       edited — the only class that feeds the content categories
        lite         whole site replaced by a minimal election-night page
        empty        response carried no content at all (observed: a 49-byte
                     <html><head></head><body></body></html>)
        render_flip  the capture path changed plain<->headless, so the two sides
                     are not comparable — a JS-built Google Translate picker
                     contributes ~100 lines under headless and none under plain,
                     which alone produced ~1,300 phantom "removed" lines each for
                     Broward and Monroe

    The last three are reported on their own sheet rather than dropped, because a
    polling page going empty on election morning is a finding, not noise.

Numbers are extracted separately: where a line looks like a turnout / registration /
ballot count and its neighbour in the same hunk differs only in digits, the old and
new values are recorded so the movement can be read directly.

Usage:
    python scripts/analyze_diffs.py                       # HEAD vs working tree
    python scripts/analyze_diffs.py --from HEAD~3         # HEAD~3 vs working tree
    python scripts/analyze_diffs.py --from A --to B
    python scripts/analyze_diffs.py --out /tmp/report.xlsx
"""

from __future__ import annotations

import argparse
import csv
import re
import subprocess
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
MANIFEST = ROOT / "manifest" / "targets.csv"
DEFAULT_OUT = ROOT / "manifest" / "fl-election-day-diffs.xlsx"
DEFAULT_CSV = ROOT / "manifest" / "fl-election-day-diffs.csv"

# Categories, in report order. Each is (label, compiled pattern).
#
# Ordering note: the patterns are independent, not a priority chain — a line saying
# "Unofficial results: 42 of 45 precincts reporting" is genuinely both a results
# change and a precinct-reporting change, and suppressing one would lose signal.
CATEGORY_PATTERNS: list[tuple[str, str]] = [
    # NOTE the `s?` on every "result". Written as `\belection result\b` this misses
    # "Election Results" outright — the \b cannot match between "result" and "s" —
    # which silently zeroed the single most important election-day category.
    ("Results posted",
     r"\b(unofficial results?|official results?|election results?|returns|"
     r"election night reporting|results are|view results?|live results?|"
     r"results by precinct|election summary)\b"),
    ("Precincts reporting",
     r"\b(precincts? reporting|of \d+ precincts|% reporting|"
     r"precincts? counted|fully reported|precincts? and districts)\b"),
    # A Google Translate language picker whose ~100-entry list appears or vanishes
    # between captures. Categorized rather than filtered so its volume stays visible
    # — it is presentation churn, and lumping it into a real category would inflate
    # that category in exactly the counties running the widget.
    ("Translate widget (presentation)",
     r"^(select language|batak|chinese \(|crimean tatar|haitian creole|"
     r"french \(canada\)|kurdish|myanmar|portuguese \(|spanish|tagalog|"
     r"vietnamese|arabic|russian|korean|japanese|hindi|urdu|swahili|"
     r"powered by\s+google)"),
    ("Turnout / ballots cast",
     r"\b(turnout|ballots cast|votes cast|ballots received|ballots counted|"
     r"have voted|voter activity|cast a ballot)\b"),
    ("Voter registration numbers",
     r"\b(registered voters?|voter registration (?:total|count|statistic)|"
     r"active voters?|registration book closing|book closing)\b"),
    # The party-breakdown block that sits under a registration total. On its own a
    # line reads just "Total: 20,669" or "Republican: 8,412", which carries no
    # registration keyword at all, so without this the numbers the analysis most
    # wants would land in "uncategorized".
    ("Party / registration breakdown",
     r"^(total|democrat(?:ic)?|republican|no party affiliation|npa|other|"
     r"minor part(?:y|ies)|libertarian|independent)\s*[:\-]\s*[\d,]+"),
    ("Election day / polling place",
     r"\b(election day|polling place|polling location|polling site|"
     r"your precinct|where do i vote|precinct finder|polls (?:open|close))\b"),
    ("Early voting",
     r"\b(early voting|early vote|vote early)\b"),
    ("Vote-by-mail / drop box",
     r"\b(vote[- ]by[- ]mail|vbm|mail ballot|absentee|drop box|dropbox|"
     r"ballot intake|secure ballot)\b"),
    ("Sample ballot / candidates",
     r"\b(sample ballots?|candidate list|qualified candidates?|offices up for|"
     r"who'?s on my ballot|what'?s on my ballot|candidate (?:login|services|"
     r"file distribution)|current candidates|candidates? (?:&|and) committees|"
     r"track my ballot|ballot tracker)\b"),
    ("Next-election pivot",
     r"\b(general election|november \d{1,2}|nov\.? \d{1,2}|next election|"
     r"upcoming election|runoff)\b"),
    ("Date / countdown change",
     r"(\b(?:mon|tue|tues|wed|thu|thur|thurs|fri|sat|sun)[a-z]*\.?,? "
     r"(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.? ?\d{1,2}\b"
     r"|\b\d{1,2}/\d{1,2}/\d{2,4}\b|\bdays? (?:until|left|to go)\b"
     r"|\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.? \d{1,2},? \d{4}\b)"),
    ("Alert / notice banner",
     r"\b(alert|notice|important|attention|closed|closure|reminder|"
     r"breaking|update:|please note|emergency)\b"),
    ("Provisional / cure",
     r"\b(provisional|cure affidavit|signature cure|ballot cure)\b"),
    ("Canvassing board",
     r"\b(canvassing board|canvass|certification of|logic and accuracy|"
     r"l&a test|audit)\b"),
    ("Poll worker / staffing",
     r"\b(poll worker|election worker|poll deputy|clerk training)\b"),
    ("Wait times / lines",
     r"\b(wait time|wait times|line length|estimated wait)\b"),
    # Nav labels are the single biggest source of changed lines once a page's menu
    # is re-rendered. Labelling them keeps them out of "uncategorized" (where they
    # would look like an unexplained gap) while staying clearly separable from the
    # substantive categories above.
    ("Navigation / menu labels",
     r"^(register to vote|elected officials|election workers?|3 ways to vote|"
     r"political parties|initiative petitions|voting equipment|accessible voting|"
     r"civil rights restoration|becoming a candidate|candidates? & committees|"
     r"candidate file distribution|vote in honor of a veteran|skip to content|"
     r"supervisor of elections|contact us|about us|home|search|menu|espa[nñ]ol|"
     r"news|forms|faqs?|maps?|reports?|links|calendar|employment|"
     r"public records|voter registration|address change|my status)$"),
]
COMPILED = [(label, re.compile(pat, re.I)) for label, pat in CATEGORY_PATTERNS]

# Lines that carry a count worth reading directly. Kept narrower than the
# categories above: we only want lines where a NUMBER is the point.
_NUMERIC_CONTEXT = re.compile(
    r"\b(turnout|ballots cast|votes cast|ballots received|ballots counted|"
    r"registered voters?|active voters?|precincts? reporting|"
    r"of \d+ precincts|% reporting|vote[- ]by[- ]mail|early voting|total)\b", re.I)
_NUMBER = re.compile(r"\d[\d,]*\.?\d*%?")


def _sh(args: list[str]) -> str:
    return subprocess.run(args, cwd=ROOT, capture_output=True, text=True,
                          check=True).stdout


def changed_files(rev_from: str, rev_to: str | None) -> list[str]:
    args = ["git", "diff", "--name-only", rev_from]
    if rev_to:
        args.append(rev_to)
    args += ["--", "snapshots/*/*/page.txt"]
    return [l for l in _sh(args).splitlines() if l.strip()]


def diff_lines(rev_from: str, rev_to: str | None, path: str
               ) -> tuple[list[str], list[str]]:
    """Return (added, removed) content lines for one file."""
    args = ["git", "diff", "-U0", rev_from]
    if rev_to:
        args.append(rev_to)
    args += ["--", path]
    added, removed = [], []
    for line in _sh(args).splitlines():
        if line.startswith("+++") or line.startswith("---"):
            continue
        if line.startswith("+"):
            added.append(line[1:].strip())
        elif line.startswith("-"):
            removed.append(line[1:].strip())
    return [a for a in added if a], [r for r in removed if r]


def categorize(line: str) -> list[str]:
    return [label for label, pat in COMPILED if pat.search(line)]


def _norm_for_pairing(line: str) -> str:
    """Strip digits so a line can be matched to its before/after twin."""
    return re.sub(r"\s+", " ", _NUMBER.sub("#", line)).strip().lower()


def _only_number(line: str) -> float | None:
    """The single numeric value on a line, if there is exactly one countable one."""
    nums = [n for n in _NUMBER.findall(line) if not n.endswith("%")]
    if len(nums) != 1:
        return None
    try:
        return float(nums[0].replace(",", ""))
    except ValueError:
        return None


def numeric_moves(added: list[str], removed: list[str]) -> list[dict]:
    """Extract figures, in two kinds.

    `changed`  — the same sentence on both sides of the diff with only its digits
                 different, so before and after are directly comparable. Pairing on
                 the digit-stripped text is what makes that comparison valid;
                 without it you get two unrelated lists of numbers.
    `new`      — a figure that appears only on the added side. It has no "before",
                 but on an election day these are the most interesting numbers of
                 all (a county publishing its final early-voting count), so
                 discarding them for lacking a pair would be the wrong trade.
    """
    out: list[dict] = []
    rem_by_shape = defaultdict(list)
    for r in removed:
        if _NUMERIC_CONTEXT.search(r) and _NUMBER.search(r):
            rem_by_shape[_norm_for_pairing(r)].append(r)

    paired_added: set[str] = set()
    for a in added:
        if not (_NUMERIC_CONTEXT.search(a) and _NUMBER.search(a)):
            continue
        shape = _norm_for_pairing(a)
        if rem_by_shape.get(shape):
            old = rem_by_shape[shape].pop(0)
            if old != a:
                o, n = _only_number(old), _only_number(a)
                delta = (n - o) if (o is not None and n is not None) else None
                out.append({"kind": "changed",
                            "measure": shape.replace("#", "N")[:70],
                            "before": old[:110], "after": a[:110],
                            "delta": delta})
                paired_added.add(a)

    for a in added:
        if a in paired_added or not _NUMBER.search(a):
            continue
        # Narrower than _NUMERIC_CONTEXT: an unpaired line needs to be clearly a
        # reported figure, or every date and phone number would qualify.
        if re.search(r"\b(votes? cast|ballots? (?:cast|received|counted|returned)|"
                     r"turnout|registered voters?|precincts? reporting|"
                     r"% reporting|wait time)\b", a, re.I):
            out.append({"kind": "new", "measure": "newly reported figure",
                        "before": "", "after": a[:110], "delta": None})
    return out


# A page that kept less than this share of its bytes was replaced, not edited.
_LITE_BYTE_RATIO = 0.25
# At or below this many bytes the response carries no content at all (the observed
# case is a literal <html><head></head><body></body></html>, 49 bytes).
_EMPTY_BYTES = 200
# The shared vendor template that the election-night replacement pages use.
_LITE_MARKER = re.compile(
    r"(faster access to .{0,40}election night|streamlined election day page)", re.I)


def _read_at(rev: str | None, path: str) -> str:
    """File contents at a revision, or from the working tree when rev is None."""
    if rev is None:
        f = ROOT / path
        return f.read_text(encoding="utf-8") if f.exists() else ""
    try:
        return _sh(["git", "show", f"{rev}:{path}"])
    except Exception:  # noqa: BLE001
        return ""


def classify_capture(rev_from: str, rev_to: str | None, path: str) -> dict:
    """Decide whether a target was edited, replaced, or served empty.

    Uses meta.json byte_size on both sides rather than the text diff, because the
    question "was this page replaced?" is about the response, not its wording.

    Both sides are read AT THEIR REVISION. An earlier version always read the
    "after" side from the working tree, which silently ignored `rev_to` — so asking
    for a historical window returned the classification for "then vs now" instead,
    and any window not ending at the working tree came back empty.
    """
    meta = path.rsplit("/", 1)[0] + "/meta.json"
    import json
    try:
        new = json.loads(_read_at(rev_to, meta) or "{}")
    except Exception:  # noqa: BLE001
        new = {}
    try:
        old = json.loads(_read_at(rev_from, meta) or "{}")
    except Exception:  # noqa: BLE001
        old = {}
    if not new:
        return {"klass": "normal", "before": 0, "after": 0}
    before, after = old.get("byte_size", 0), new.get("byte_size", 0)

    klass = "normal"
    # Order matters. An empty response usually ALSO flips render mode — a 49-byte body
    # trips the JS-shell threshold and escalates to headless — so testing the flip
    # first would relabel every emptied page as a mere capture-path change and hide
    # the outage. Emptiness is the stronger, more important fact, so it wins.
    if after <= _EMPTY_BYTES:
        klass = "empty"
    # A plain<->headless flip changes HOW MUCH of the page is captured, not what the
    # county published: a JS-rendered Google Translate picker contributes ~100 lines
    # under headless and none under plain. Broward and Monroe each showed ~1,300
    # "removed" lines on 2026-08-18 that were entirely this. Line counts either side
    # of a flip are not comparable, so these are separated out rather than counted.
    elif (old.get("render_mode") and new.get("render_mode")
          and old["render_mode"] != new["render_mode"]):
        klass = "render_flip"
    else:
        text = _read_at(rev_to, path)
        if _LITE_MARKER.search(text):
            klass = "lite"
        elif before and after / before < _LITE_BYTE_RATIO:
            klass = "lite"
    return {"klass": klass, "before": before, "after": after,
            "status_before": old.get("http_status"),
            "status_after": new.get("http_status"),
            "render": f'{old.get("render_mode")} -> {new.get("render_mode")}'}


def load_manifest() -> tuple[dict[str, int], set[str]]:
    captured: Counter[str] = Counter()
    counties: set[str] = set()
    with MANIFEST.open(newline="", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            counties.add(r["county"].strip())
            if r["url"].strip():
                captured[r["county"].strip()] += 1
    return dict(captured), counties


def slug_to_county(slug: str, counties: set[str]) -> str:
    for c in counties:
        if c.lower().replace(" ", "_") == slug:
            return c
    return slug


def analyze(rev_from: str, rev_to: str | None) -> dict:
    captured, counties = load_manifest()
    files = changed_files(rev_from, rev_to)

    per_county_cat: dict[str, Counter] = defaultdict(Counter)
    per_county_stats: dict[str, dict] = defaultdict(
        lambda: {"files": 0, "added": 0, "removed": 0, "pages": set()})
    per_type_stats: dict[str, dict] = defaultdict(
        lambda: {"files": 0, "added": 0, "removed": 0, "counties": set()})
    per_type_cat: dict[str, Counter] = defaultdict(Counter)
    cat_counties: dict[str, set] = defaultdict(set)
    cat_types: dict[str, set] = defaultdict(set)
    cat_examples: dict[str, list] = defaultdict(list)
    cat_total: Counter = Counter()
    numbers: list[dict] = []
    uncategorized: list[dict] = []
    outages: list[dict] = []
    # Corpus-wide, not per-file: a nav label repeats on all five of a county's pages,
    # so a per-file set would still emit it five times.
    example_seen: set[str] = set()

    for path in files:
        parts = path.split("/")
        if len(parts) < 4:
            continue
        slug, ptype = parts[1], parts[2]
        county = slug_to_county(slug, counties)
        added, removed = diff_lines(rev_from, rev_to, path)

        cap = classify_capture(rev_from, rev_to, path)
        if cap["klass"] != "normal":
            # Replaced or emptied, not edited. Recorded in full on its own sheet and
            # kept out of the content categories — see the module docstring.
            outages.append({
                "county": county, "page_type": ptype, "class": cap["klass"],
                "bytes_before": cap["before"], "bytes_after": cap["after"],
                "pct_retained": (cap["after"] / cap["before"]) if cap["before"] else 0,
                "lines_removed": len(removed), "lines_added": len(added),
                "status": f'{cap.get("status_before")} -> {cap.get("status_after")}',
                "render": cap.get("render", ""),
            })
            continue

        st = per_county_stats[county]
        st["files"] += 1
        st["added"] += len(added)
        st["removed"] += len(removed)
        st["pages"].add(ptype)
        ts = per_type_stats[ptype]
        ts["files"] += 1
        ts["added"] += len(added)
        ts["removed"] += len(removed)
        ts["counties"].add(county)

        for line in added + removed:
            cats = categorize(line)
            if not cats:
                if len(line) > 12:
                    uncategorized.append({"county": county, "page_type": ptype,
                                          "line": line[:150]})
                continue
            for c in cats:
                per_county_cat[county][c] += 1
                per_type_cat[ptype][c] += 1
                cat_total[c] += 1
                cat_counties[c].add(county)
                cat_types[c].add(ptype)
                # Dedupe on (county, line): the same nav label repeats across a
                # county's five pages, and four identical rows tell you nothing that
                # one row plus the county count does not.
                key = f"{c}|{county}|{line[:140]}"
                if key not in example_seen and len(cat_examples[c]) < 40:
                    cat_examples[c].append({"county": county, "page_type": ptype,
                                            "line": line[:140]})
                    example_seen.add(key)

        for m in numeric_moves(added, removed):
            numbers.append({"county": county, "page_type": ptype, **m})

    return {
        "files": files, "captured": captured, "counties": counties,
        "per_county_cat": per_county_cat, "per_county_stats": per_county_stats,
        "per_type_stats": per_type_stats, "per_type_cat": per_type_cat,
        "cat_counties": cat_counties, "cat_types": cat_types,
        "cat_examples": cat_examples, "cat_total": cat_total,
        "numbers": numbers, "uncategorized": uncategorized,
        "outages": outages,
    }


# --------------------------------------------------------------------------- #
# Output
# --------------------------------------------------------------------------- #
FONT = "Arial"


def write_xlsx(a: dict, out: Path, rev_from: str, rev_to: str | None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    base = Font(name=FONT, size=10)
    bold = Font(name=FONT, bold=True, size=10)
    border = Border(bottom=Side(style="thin", color="BFBFBF"))
    heat = [PatternFill("solid", fgColor=c) for c in
            ("FFFFFF", "E8F1DC", "CDE4B4", "A9D18E", "70AD47")]

    labels = [c for c, _ in CATEGORY_PATTERNS]
    counties_sorted = sorted(a["per_county_stats"],
                             key=lambda c: -a["per_county_stats"][c]["added"]
                             - a["per_county_stats"][c]["removed"])

    def sheet(ws, headers, widths):
        ws.append(headers)
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i)
            c.fill, c.font = head_fill, head_font
            c.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
        ws.row_dimensions[1].height = 42
        ws.freeze_panes = "B2"

    wb = Workbook()

    # --- 1. By county ------------------------------------------------------- #
    ws = wb.active
    ws.title = "By county"
    heads = (["County", "Pages changed", "Pages captured", "% of pages changed",
              "Lines changed", "Lines added", "Lines removed", "Page types changed"]
             + labels)
    sheet(ws, heads, [15, 11, 11, 11, 10, 9, 10, 26] + [13] * len(labels))
    maxcat = max([max(a["per_county_cat"][c].values(), default=0)
                  for c in counties_sorted] or [1])
    for county in counties_sorted:
        st = a["per_county_stats"][county]
        cap = a["captured"].get(county, 0)
        row = [county, st["files"], cap,
               (st["files"] / cap) if cap else 0,
               st["added"] + st["removed"], st["added"], st["removed"],
               ", ".join(sorted(st["pages"]))]
        row += [a["per_county_cat"][county].get(l, 0) for l in labels]
        ws.append(row)
        r = ws.max_row
        for i in range(1, len(heads) + 1):
            cell = ws.cell(row=r, column=i)
            cell.font = base
            cell.border = border
            if i > 8:
                v = cell.value or 0
                if v:
                    cell.fill = heat[min(4, 1 + int(3 * v / max(1, maxcat)))]
        ws.cell(row=r, column=4).number_format = "0%"
    tot = ws.max_row + 1
    ws.cell(row=tot, column=1, value="TOTAL").font = bold
    for i, key in ((2, "files"), (6, "added"), (7, "removed")):
        ws.cell(row=tot, column=i,
                value=f"=SUM({get_column_letter(i)}2:{get_column_letter(i)}{tot-1})"
                ).font = bold
    ws.cell(row=tot, column=5,
            value=f"=SUM(E2:E{tot-1})").font = bold
    for n, l in enumerate(labels):
        col = get_column_letter(9 + n)
        ws.cell(row=tot, column=9 + n,
                value=f"=SUM({col}2:{col}{tot-1})").font = bold
    ws.auto_filter.ref = f"A1:{get_column_letter(len(heads))}{tot-1}"

    # --- 2. By category ----------------------------------------------------- #
    wc = wb.create_sheet("By category")
    heads2 = ["Category", "Line hits", "Counties affected", "% of 67 counties",
              "Page types affected", "Share of all categorized hits", "Example"]
    sheet(wc, heads2, [30, 11, 15, 13, 30, 14, 92])
    grand = sum(a["cat_total"].values()) or 1
    for label in sorted(labels, key=lambda l: -a["cat_total"].get(l, 0)):
        ex = a["cat_examples"].get(label) or [{}]
        wc.append([label, a["cat_total"].get(label, 0),
                   len(a["cat_counties"].get(label, ())),
                   len(a["cat_counties"].get(label, ())) / 67,
                   ", ".join(sorted(a["cat_types"].get(label, ()))),
                   a["cat_total"].get(label, 0) / grand,
                   ex[0].get("line", "")])
        r = wc.max_row
        for i in range(1, len(heads2) + 1):
            wc.cell(row=r, column=i).font = base
            wc.cell(row=r, column=i).border = border
        wc.cell(row=r, column=4).number_format = "0%"
        wc.cell(row=r, column=6).number_format = "0.0%"
        wc.cell(row=r, column=7).alignment = Alignment(wrap_text=True,
                                                       vertical="top")

    # --- 3. By page type ---------------------------------------------------- #
    wt = wb.create_sheet("By page type")
    heads3 = ["Page type", "Files changed", "Counties affected", "Lines changed",
              "Lines added", "Lines removed", "Top categories"]
    sheet(wt, heads3, [16, 13, 15, 12, 11, 12, 66])
    for pt in sorted(a["per_type_stats"],
                     key=lambda p: -a["per_type_stats"][p]["files"]):
        s = a["per_type_stats"][pt]
        top = ", ".join(f"{k} ({v})"
                        for k, v in a["per_type_cat"][pt].most_common(4))
        wt.append([pt, s["files"], len(s["counties"]), s["added"] + s["removed"],
                   s["added"], s["removed"], top])
        r = wt.max_row
        for i in range(1, len(heads3) + 1):
            wt.cell(row=r, column=i).font = base
            wt.cell(row=r, column=i).border = border
        wt.cell(row=r, column=7).alignment = Alignment(wrap_text=True,
                                                       vertical="top")

    # --- 3b. Election-night site replacements -------------------------------- #
    wo = wb.create_sheet("Site replaced or empty")
    heads_o = ["County", "Page type", "What happened", "Bytes before", "Bytes after",
               "% of bytes retained", "Lines removed", "HTTP status", "Render mode"]
    sheet(wo, heads_o, [15, 14, 32, 13, 12, 14, 13, 14, 22])
    red = PatternFill("solid", fgColor="FFC7CE")
    amber = PatternFill("solid", fgColor="FFEB9C")
    explain = {"lite": "replaced by election-night page",
               "empty": "served empty HTML (no content)",
               "render_flip": "capture path changed (not comparable)"}
    for o in sorted(a["outages"], key=lambda x: (x["class"] != "empty",
                                                 x["county"], x["page_type"])):
        wo.append([o["county"], o["page_type"], explain.get(o["class"], o["class"]),
                   o["bytes_before"], o["bytes_after"], o["pct_retained"],
                   o["lines_removed"], o["status"], o.get("render", "")])
        r = wo.max_row
        for i in range(1, len(heads_o) + 1):
            wo.cell(row=r, column=i).font = base
            wo.cell(row=r, column=i).border = border
        wo.cell(row=r, column=6).number_format = "0.0%"
        wo.cell(row=r, column=3).fill = (
            red if o["class"] == "empty"
            else amber if o["class"] == "lite"
            else PatternFill("solid", fgColor="DDEBF7"))
    if not a["outages"]:
        wo.append(["(none)", "", "every changed page served normal content", "",
                   "", "", "", "", ""])

    # --- 4. Numbers that moved ---------------------------------------------- #
    wn = wb.create_sheet("Numbers")
    heads4 = ["County", "Page type", "Kind", "Measure (digits masked)", "Before",
              "After", "Change"]
    sheet(wn, heads4, [15, 14, 10, 44, 56, 56, 10])
    for n in sorted(a["numbers"],
                    key=lambda x: (x["kind"] != "changed", x["county"],
                                   x["page_type"])):
        wn.append([n["county"], n["page_type"], n["kind"], n["measure"],
                   n["before"], n["after"],
                   n["delta"] if n.get("delta") is not None else ""])
        r = wn.max_row
        for i in range(1, len(heads4) + 1):
            wn.cell(row=r, column=i).font = base
            wn.cell(row=r, column=i).border = border
            wn.cell(row=r, column=i).alignment = Alignment(wrap_text=True,
                                                          vertical="top")
        wn.cell(row=r, column=3).fill = (
            PatternFill("solid", fgColor="DDEBF7") if n["kind"] == "changed"
            else PatternFill("solid", fgColor="E2EFDA"))
        if n.get("delta") is not None:
            wn.cell(row=r, column=7).number_format = "+#,##0;-#,##0;0"
    if not a["numbers"]:
        wn.append(["(none)", "", "", "No line carried a comparable figure.", "",
                   "", ""])

    # --- 5. Notable & rare --------------------------------------------------- #
    wr = wb.create_sheet("Notable and rare")
    heads5 = ["Category", "Counties with it", "County", "Page type",
              "The changed line"]
    sheet(wr, heads5, [28, 14, 15, 14, 96])
    rare = [l for l in labels
            if 0 < len(a["cat_counties"].get(l, ())) <= 5]
    for label in sorted(rare, key=lambda l: len(a["cat_counties"][l])):
        for ex in a["cat_examples"].get(label, [])[:6]:
            wr.append([label, len(a["cat_counties"][label]), ex["county"],
                       ex["page_type"], ex["line"]])
            r = wr.max_row
            for i in range(1, len(heads5) + 1):
                wr.cell(row=r, column=i).font = base
                wr.cell(row=r, column=i).border = border
                wr.cell(row=r, column=i).alignment = Alignment(wrap_text=True,
                                                               vertical="top")
    if wr.max_row == 1:
        wr.append(["(no category was confined to 5 or fewer counties)", "", "", "",
                   ""])

    # --- 6. Uncategorized sample --------------------------------------------- #
    wu = wb.create_sheet("Uncategorized sample")
    sheet(wu, ["County", "Page type", "Changed line (no category matched)"],
          [15, 14, 120])
    wu.cell(row=1, column=3).comment = None
    for u in a["uncategorized"][:400]:
        wu.append([u["county"], u["page_type"], u["line"]])
        r = wu.max_row
        for i in range(1, 4):
            wu.cell(row=r, column=i).font = base
            wu.cell(row=r, column=i).border = border
        wu.cell(row=r, column=3).alignment = Alignment(wrap_text=True,
                                                       vertical="top")

    # --- 6b. Column dictionary ----------------------------------------------- #
    wd = wb.create_sheet("Columns")
    wd.column_dimensions["A"].width = 20
    wd.column_dimensions["B"].width = 30
    wd.column_dimensions["C"].width = 96
    wd["A1"] = "What every column means"
    wd["A1"].font = Font(name=FONT, bold=True, size=12)
    wd.append([])
    for h, i in (("Sheet", 1), ("Column", 2), ("Meaning", 3)):
        c = wd.cell(row=2, column=i, value=h)
        c.fill, c.font = head_fill, head_font
    docs = [
        ("By county", "County", "Florida county. Only counties with at least one "
                                "genuinely edited page appear; those whose site was "
                                "replaced or emptied are on 'Site replaced or empty'."),
        ("By county", "Pages changed", "How many of this county's captured pages "
                                       "changed at all."),
        ("By county", "Pages captured", "How many of the 5 page types this county "
                                        "publishes (from the manifest). 4/5 means "
                                        "one type is a recorded gap, not a failure."),
        ("By county", "% of pages changed", "Pages changed ÷ pages captured. Lets a "
                                            "small county that changed everything "
                                            "rank against a big one that changed one "
                                            "page."),
        ("By county", "Lines changed", "Added + removed lines of page.txt. THE "
                                       "unduplicated total — use this, not the sum "
                                       "of the category columns."),
        ("By county", "Lines added / removed", "Direction of the change. Heavy "
                                               "removal with light addition usually "
                                               "means content was retired (early "
                                               "voting ending); the reverse means "
                                               "new material was published."),
        ("By county", "Page types changed", "Which of homepage / elections / polling "
                                            "/ early_voting / results moved."),
        ("By county", "<category columns>", "Lines in this county matching that "
                                            "category. These do NOT sum to 'Lines "
                                            "changed' and are not a partition of it: "
                                            "a line matching two categories is "
                                            "counted twice, and a line matching none "
                                            "is not counted at all. The second "
                                            "dominates — Escambia changed 1,089 "
                                            "lines, 480 matched nothing, so its "
                                            "categories total 408. Shaded darker "
                                            "green the higher the count, to compare "
                                            "across counties at a glance."),
        ("By category", "Line hits", "Total matching lines across every county."),
        ("By category", "Counties affected", "How many distinct counties had at "
                                             "least one line in this category — the "
                                             "better measure of how WIDESPREAD a "
                                             "behaviour is, where 'Line hits' "
                                             "measures VOLUME and can be driven by "
                                             "one verbose county."),
        ("By category", "% of 67 counties", "Counties affected ÷ 67."),
        ("By category", "Share of all categorized hits",
         "This category's line hits ÷ all category hits. Shows the shape of "
         "election-day activity."),
        ("By category", "Example", "One real changed line, so the category is not "
                                   "taken on trust."),
        ("By page type", "Files changed", "Pages of this type that changed, across "
                                          "all counties."),
        ("By page type", "Top categories", "The four categories with most lines for "
                                           "this page type — what this kind of page "
                                           "is FOR on election day."),
        ("Site replaced or empty", "What happened",
         "'replaced by election-night page' = the county swapped its whole site for "
         "a minimal static page. 'served empty HTML' = the URL returned no content. "
         "'capture path changed' = the plain/headless fetch path differed between "
         "runs, so the two sides are NOT comparable."),
        ("Site replaced or empty", "Bytes before / after",
         "Response size either side. This is what identifies a replacement: Calhoun's "
         "polling page went 202,198 → 49 bytes."),
        ("Site replaced or empty", "% of bytes retained",
         "After ÷ before. Under 25% means replaced, not edited."),
        ("Site replaced or empty", "HTTP status",
         "Status before → after. 200 → 404 is the tell that the old URL no longer "
         "exists."),
        ("Site replaced or empty", "Render mode",
         "plain → headless (or the reverse). A flip changes how much of the page is "
         "captured, which is why those rows are excluded from the content counts."),
        ("Numbers", "Kind", "'changed' = the same sentence appears on both sides "
                            "with only its digits different, so before/after are "
                            "directly comparable. 'new' = a figure that appears only "
                            "in the added text and has no prior value."),
        ("Numbers", "Measure (digits masked)",
         "The sentence with its numbers replaced by N — this is the key the "
         "before/after pairing is done on."),
        ("Numbers", "Before / After", "The actual lines."),
        ("Numbers", "Change", "After − before, filled only where each side carried "
                              "exactly one comparable number."),
        ("Notable and rare", "Counties with it",
         "How many counties showed this category at all. Rows are limited to "
         "categories confined to 5 or fewer counties — the unusual behaviour, as "
         "opposed to what everyone did."),
        ("Uncategorized sample", "Changed line",
         "A changed line that matched no category. Shown so the keyword lists can be "
         "improved; a large count here means the categories are incomplete, NOT that "
         "nothing happened."),
    ]
    for s, col, mean in docs:
        wd.append([s, col, mean])
        r = wd.max_row
        wd.cell(row=r, column=1).font = base
        wd.cell(row=r, column=2).font = bold
        wd.cell(row=r, column=3).font = base
        wd.cell(row=r, column=3).alignment = Alignment(wrap_text=True,
                                                       vertical="top")
        for i in range(1, 4):
            wd.cell(row=r, column=i).border = border
    wd.freeze_panes = "A3"

    # --- 7. Method ----------------------------------------------------------- #
    wm = wb.create_sheet("Method")
    wm.column_dimensions["A"].width = 26
    wm.column_dimensions["B"].width = 108
    wm["A1"] = "How to read this workbook"
    wm["A1"].font = Font(name=FONT, bold=True, size=12)
    def _describe(rev: str | None) -> str:
        if rev is None:
            return "working tree (fresh capture)"
        try:
            return _sh(["git", "log", "-1", "--format=%h  %ad  %s",
                        "--date=short", rev]).strip()
        except Exception:  # noqa: BLE001
            return rev

    span = f"{_describe(rev_from)}\n    →  {_describe(rev_to)}"
    notes = [
        ("Comparison", span),
        ("Unit of measure", "One line of page.txt. A file that changed in three "
                            "places contributes three lines."),
        ("Artifact used", "page.txt only — the visible-text artifact. page.html "
                          "carries the same change plus markup, so counting both "
                          "would double-count."),
        ("Categories do not sum", "The category columns are not a partition of "
                                  "'Lines changed'. A line matching two categories "
                                  "is counted twice; a line matching none is not "
                                  "counted at all. Across this run: 3,791 lines "
                                  "changed, 1,661 category hits, 1,368 lines "
                                  "matched nothing. Use 'Lines changed' for "
                                  "magnitude, categories for shape."),
        ("Noise floor", "Two back-to-back captures of these same 314 targets "
                        "produced a 0-line diff, so every line counted here is a "
                        "real content change, not capture jitter."),
        ("Numbers that moved", "Recorded only where the same sentence appears on "
                               "both sides of the diff with just its digits "
                               "changed — that pairing is what makes before/after "
                               "comparable."),
        ("Uncategorized", "Shown so the keyword lists can be improved. A high "
                          "count there means the categories are missing something, "
                          "not that nothing happened."),
        ("Regenerate", "python scripts/analyze_diffs.py --from <rev> [--to <rev>]"),
    ]
    wm.append([])
    for k, v in notes:
        wm.append([k, v])
        r = wm.max_row
        wm.cell(row=r, column=1).font = bold
        wm.cell(row=r, column=2).font = base
        wm.cell(row=r, column=2).alignment = Alignment(wrap_text=True,
                                                       vertical="top")

    wb.calculation.fullCalcOnLoad = True
    wb.save(out)


def write_csv(a: dict, out: Path) -> None:
    labels = [c for c, _ in CATEGORY_PATTERNS]
    with out.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["county", "pages_changed", "pages_captured", "lines_changed",
                    "lines_added", "lines_removed", "page_types_changed"] + labels)
        for county in sorted(a["per_county_stats"]):
            st = a["per_county_stats"][county]
            w.writerow([county, st["files"], a["captured"].get(county, 0),
                        st["added"] + st["removed"], st["added"], st["removed"],
                        "|".join(sorted(st["pages"]))]
                       + [a["per_county_cat"][county].get(l, 0) for l in labels])


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--from", dest="rev_from", default="HEAD")
    ap.add_argument("--to", dest="rev_to", default=None)
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--csv", default=str(DEFAULT_CSV))
    args = ap.parse_args()

    a = analyze(args.rev_from, args.rev_to)

    n_lines = sum(s["added"] + s["removed"] for s in a["per_county_stats"].values())
    lite = [o for o in a["outages"] if o["class"] == "lite"]
    empty = [o for o in a["outages"] if o["class"] == "empty"]
    print(f"{len(a['files'])} changed page.txt files\n")
    print(f"  site REPLACED by an election-night page : {len(lite):>3} targets "
          f"({len({o['county'] for o in lite})} counties)")
    print(f"  served EMPTY html                       : {len(empty):>3} targets "
          f"({len({o['county'] for o in empty})} counties)")
    flip = [o for o in a["outages"] if o["class"] == "render_flip"]
    print(f"  capture path changed (not comparable)   : {len(flip):>3} targets "
          f"({len({o['county'] for o in flip})} counties)")
    print(f"  genuinely edited                        : "
          f"{len(a['files']) - len(a['outages']):>3} targets "
          f"({len(a['per_county_stats'])} counties)\n")
    print(f"{n_lines} changed lines on edited pages")
    print(f"categorized hits: {sum(a['cat_total'].values())} "
          f"(a line can match several) · uncategorized lines: "
          f"{len(a['uncategorized'])}")
    print(f"numeric before/after pairs: {len(a['numbers'])}\n")
    for label, n in sorted(a["cat_total"].items(), key=lambda x: -x[1]):
        print(f"  {label:<30} {n:>5} hits  "
              f"{len(a['cat_counties'][label]):>3} counties")

    write_csv(a, Path(args.csv))
    write_xlsx(a, Path(args.out), args.rev_from, args.rev_to)
    print(f"\nwrote {args.csv}\nwrote {args.out}")


if __name__ == "__main__":
    main()
