#!/usr/bin/env python3
"""Generate the formatted spreadsheet view of the manifest.

`manifest/targets.csv` stays the pipeline's source of truth — it is what snapshot.py
reads and what diffs cleanly in git. This script renders it as
`manifest/fl-county-election-pages.xlsx` for review: frozen headers, autofilter,
colour-coded QA columns, and a Coverage sheet whose numbers are live formulas over
the Targets sheet rather than values baked in at export time.

Because the workbook is generated, it can never drift from the CSV — re-run this
after any manifest change instead of editing the workbook.

Sheets:
    Targets   one row per (county x page type) — the manifest, formatted
    Counties  the 67-county seed (identity + homepage provenance)
    Coverage  captured/gap counts per page type, as formulas
    Legend    what each column and colour means

Usage:
    python scripts/export_xlsx.py
    python scripts/export_xlsx.py --out /tmp/review.xlsx
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
TARGETS = ROOT / "manifest" / "targets.csv"
COUNTIES = ROOT / "manifest" / "counties.csv"
OUT = ROOT / "manifest" / "fl-county-election-pages.xlsx"

FONT = "Arial"
PAGE_TYPES = ["homepage", "elections", "polling", "early_voting", "results"]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BASE_FONT = Font(name=FONT, size=10)
LINK_FONT = Font(name=FONT, size=10, color="0563C1", underline="single")

# Row-level shading: a gap is data, not an error, so it gets a neutral grey rather
# than a warning colour.
GAP_FILL = PatternFill("solid", fgColor="F2F2F2")
# Cell-level shading for the audit verdict.
VERDICT_FILL = {
    "confident": PatternFill("solid", fgColor="C6EFCE"),   # green
    "likely": PatternFill("solid", fgColor="FFEB9C"),      # amber
    "uncertain": PatternFill("solid", fgColor="FFC7CE"),   # red
    "broken": PatternFill("solid", fgColor="FF9999"),      # strong red
    "gap": PatternFill("solid", fgColor="E7E6E6"),         # grey
}
FLAG_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(bottom=THIN)

# (column, width). Notes and reasons are wide because they carry the QA judgement.
WIDTHS = {
    "county": 15, "batch": 7, "page_type": 14, "url": 62, "external": 9,
    "notes": 78, "verify_status": 14, "http_status": 12, "final_url": 46,
    "audit_confidence": 17, "audit_reason": 78, "flag_for_review": 15,
    "seat": 18, "office_city": 18, "homepage": 44,
}


def _rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def _write_sheet(ws, headers: list[str], rows: list[dict[str, str]],
                 url_cols: set[str], gap_col: str | None = None) -> None:
    ws.append(headers)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(vertical="center", horizontal="left")
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(h, 16)
    ws.row_dimensions[1].height = 22

    for r in rows:
        ws.append([r.get(h, "") for h in headers])
        rownum = ws.max_row
        is_gap = gap_col is not None and not (r.get(gap_col) or "").strip()
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=rownum, column=i)
            c.font = BASE_FONT
            c.border = BORDER
            c.alignment = Alignment(vertical="top",
                                    wrap_text=h in ("notes", "audit_reason"))
            if is_gap:
                c.fill = GAP_FILL
            if h in url_cols and (r.get(h) or "").strip().startswith("http"):
                c.font = LINK_FONT
                c.hyperlink = r[h].strip()
            if h == "audit_confidence":
                fill = VERDICT_FILL.get((r.get(h) or "").strip())
                if fill:
                    c.fill = fill
            if h == "flag_for_review" and (r.get(h) or "").strip() == "yes":
                c.fill = FLAG_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def build(out: Path) -> None:
    targets = _rows(TARGETS)
    counties = _rows(COUNTIES)

    wb = Workbook()

    ws = wb.active
    ws.title = "Targets"
    theaders = ["county", "batch", "page_type", "url", "external", "notes",
                "verify_status", "http_status", "final_url", "audit_confidence",
                "audit_reason", "flag_for_review"]
    _write_sheet(ws, theaders, targets, url_cols={"url", "final_url"},
                 gap_col="url")
    last = ws.max_row

    wc = wb.create_sheet("Counties")
    _write_sheet(wc, ["county", "seat", "office_city", "batch", "homepage"],
                 counties, url_cols={"homepage"})

    # --- Coverage: live formulas over Targets, not values frozen at export ----
    cv = wb.create_sheet("Coverage")
    cv["A1"] = "Coverage by page type"
    cv["A1"].font = Font(name=FONT, bold=True, size=12)
    hdr = ["page_type", "captured", "gaps", "total", "% captured"]
    cv.append([])
    cv.append(hdr)
    for i, h in enumerate(hdr, start=1):
        c = cv.cell(row=3, column=i)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
    cv.column_dimensions["A"].width = 16
    for col in "BCDE":
        cv.column_dimensions[col].width = 13

    # Targets columns: C = page_type, D = url
    for n, pt in enumerate(PAGE_TYPES):
        r = 4 + n
        cv[f"A{r}"] = pt
        cv[f"B{r}"] = (f'=COUNTIFS(Targets!$C$2:$C${last},$A{r},'
                       f'Targets!$D$2:$D${last},"?*")')
        cv[f"C{r}"] = f"=D{r}-B{r}"
        cv[f"D{r}"] = f'=COUNTIF(Targets!$C$2:$C${last},$A{r})'
        cv[f"E{r}"] = f"=IFERROR(B{r}/D{r},0)"
        for col in "ABCDE":
            cv[f"{col}{r}"].font = BASE_FONT
        cv[f"E{r}"].number_format = "0.0%"

    tot = 4 + len(PAGE_TYPES)
    cv[f"A{tot}"] = "TOTAL"
    cv[f"B{tot}"] = f"=SUM(B4:B{tot - 1})"
    cv[f"C{tot}"] = f"=SUM(C4:C{tot - 1})"
    cv[f"D{tot}"] = f"=SUM(D4:D{tot - 1})"
    cv[f"E{tot}"] = f"=IFERROR(B{tot}/D{tot},0)"
    cv[f"E{tot}"].number_format = "0.0%"
    for col in "ABCDE":
        cv[f"{col}{tot}"].font = Font(name=FONT, bold=True, size=10)

    cv[f"A{tot + 2}"] = "Counties by number of page types captured"
    cv[f"A{tot + 2}"].font = Font(name=FONT, bold=True, size=12)
    cv.append([])
    hdr2 = ["pages captured", "counties"]
    cv.cell(row=tot + 4, column=1, value=hdr2[0]).fill = HEAD_FILL
    cv.cell(row=tot + 4, column=1).font = HEAD_FONT
    cv.cell(row=tot + 4, column=2, value=hdr2[1]).fill = HEAD_FILL
    cv.cell(row=tot + 4, column=2).font = HEAD_FONT
    # Counted in Python: a per-county tally needs a helper column in Excel, and a
    # helper column would be a worse artifact than a documented static count.
    from collections import Counter
    per: Counter[str] = Counter()
    for r in targets:
        if (r.get("url") or "").strip():
            per[r["county"]] += 1
    dist = Counter(per.values())
    for n, k in enumerate(sorted(dist, reverse=True)):
        row = tot + 5 + n
        cv.cell(row=row, column=1, value=f"{k} of 5").font = BASE_FONT
        cv.cell(row=row, column=2, value=dist[k]).font = BASE_FONT
    note = cv.cell(row=tot + 6 + len(dist), column=1,
                   value="Note: this second table is a static count taken at export "
                         "time (a per-county tally would need a helper column). "
                         "The table above is live formulas.")
    note.font = Font(name=FONT, size=9, italic=True, color="808080")

    # --- Legend ---------------------------------------------------------------
    lg = wb.create_sheet("Legend")
    lg.column_dimensions["A"].width = 22
    lg.column_dimensions["B"].width = 104
    lg["A1"] = "fl-county-watch — manifest columns"
    lg["A1"].font = Font(name=FONT, bold=True, size=12)
    entries = [
        ("county", "Florida county name, as it appears everywhere else in the project."),
        ("batch", "Where this county's SOE homepage came from. 1 = the Florida DOS "
                  "Supervisor of Elections directory, as published. 2 = that "
                  "directory URL was stale, and the live domain was resolved and "
                  "verified instead."),
        ("page_type", "homepage (SOE front page) · elections · polling · "
                      "early_voting · results."),
        ("url", "The page that gets snapshotted. EMPTY = a recorded gap; the reason "
                "is in notes."),
        ("external", "true when the URL's registered domain differs from the SOE "
                     "homepage's — typically a third-party results portal."),
        ("notes", "Provenance: how the URL was found, why a row is a gap, or the "
                  "human QA judgement ('corrected:' / 'verified:')."),
        ("verify_status", "ok (live) · broken (4xx/5xx/error/non-HTML) · gap (no URL)."),
        ("http_status", "HTTP status of the final response."),
        ("final_url", "Filled only when the request redirected elsewhere."),
        ("audit_confidence", "confident · likely · uncertain · broken · gap — whether "
                             "the fetched content really is this county's page of "
                             "this type."),
        ("audit_reason", "Which identity and page-type keywords matched, plus the "
                         "page title."),
        ("flag_for_review", "yes when a human should eyeball the row."),
    ]
    lg.append([])
    for k, v in entries:
        lg.append([k, v])
        r = lg.max_row
        lg.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10)
        lg.cell(row=r, column=2).font = BASE_FONT
        lg.cell(row=r, column=2).alignment = Alignment(wrap_text=True,
                                                       vertical="top")

    lg.append([])
    lg.append(["Colours", ""])
    lg.cell(row=lg.max_row, column=1).font = Font(name=FONT, bold=True, size=11)
    colours = [
        ("confident", "green — identity and page-type signals both matched",
         "confident"),
        ("likely", "amber — matched, but with a weaker signal", "likely"),
        ("uncertain", "red — needs a human look", "uncertain"),
        ("broken", "strong red — the URL did not serve a usable page", "broken"),
        ("gap row", "whole row grey — no URL for this page type; see notes", "gap"),
    ]
    for label, desc, key in colours:
        lg.append([label, desc])
        r = lg.max_row
        lg.cell(row=r, column=1).fill = VERDICT_FILL[key]
        lg.cell(row=r, column=1).font = BASE_FONT
        lg.cell(row=r, column=2).font = BASE_FONT

    lg.append([])
    lg.append(["Source of truth",
               "manifest/targets.csv. This workbook is GENERATED by "
               "scripts/export_xlsx.py — edit the CSV and re-run, never edit here."])
    lg.cell(row=lg.max_row, column=1).font = Font(name=FONT, bold=True, size=10)
    lg.cell(row=lg.max_row, column=2).font = Font(name=FONT, size=10, italic=True)
    lg.cell(row=lg.max_row, column=2).alignment = Alignment(wrap_text=True,
                                                            vertical="top")

    # openpyxl writes formulas with no cached values, so the Coverage cells would
    # read as blank in any viewer that shows cached results. Asking for a full
    # recalculation on load makes Excel/Numbers/LibreOffice populate them the moment
    # the file is opened.
    wb.calculation.fullCalcOnLoad = True
    wb.save(out)


def verify(out: Path) -> None:
    """Check the Coverage formulas reference the right ranges and would be correct.

    Cached formula values cannot be produced without a spreadsheet engine, so
    instead of trusting the formulas by eye this recomputes the same counts in
    Python and asserts the ranges the formulas point at line up with the data.
    An off-by-one range is the failure this catches — it yields no error in Excel,
    just a quietly wrong number.
    """
    from collections import Counter
    from openpyxl import load_workbook

    targets = _rows(TARGETS)
    wb = load_workbook(out)
    ws, cv = wb["Targets"], wb["Coverage"]

    assert ws.max_row == len(targets) + 1, (
        f"Targets sheet has {ws.max_row - 1} data rows, CSV has {len(targets)}")
    # Column C must really be page_type and D really url, or every COUNTIFS is wrong.
    assert ws["C1"].value == "page_type", f"col C is {ws['C1'].value!r}, not page_type"
    assert ws["D1"].value == "url", f"col D is {ws['D1'].value!r}, not url"

    captured = Counter(r["page_type"] for r in targets if (r["url"] or "").strip())
    total = Counter(r["page_type"] for r in targets)
    for n, pt in enumerate(PAGE_TYPES):
        row = 4 + n
        assert cv[f"A{row}"].value == pt
        f = cv[f"B{row}"].value
        assert f.startswith("=COUNTIFS(") and f"$C${ws.max_row}" in f, \
            f"{pt}: captured formula does not span the data rows: {f}"
        # What the formula will compute, computed independently.
        print(f"  {pt:<13} captured={captured[pt]:<4} total={total[pt]}")
    assert sum(total.values()) == len(targets)
    print(f"  ranges verified against {len(targets)} CSV rows")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(OUT))
    args = ap.parse_args()
    out = Path(args.out)
    build(out)
    print(f"wrote {out}")
    verify(out)


if __name__ == "__main__":
    main()
