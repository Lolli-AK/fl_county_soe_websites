#!/usr/bin/env python3
"""Turn the diff analysis into a readable per-county digest.

`analyze_diffs.py` produces a 26-column matrix of line counts. That is the right
shape for a machine and the wrong shape for a person: nobody has intuition for
"1,089 lines changed", and 17 category columns is more than anyone can read across.

This renders the same underlying data as **one plain-English row per county**, for
all 67 — including the counties that changed nothing and the ones that replaced their
site, which the matrix hid on other sheets. The 17 categories are kept in full; they
just stop being columns and become (a) the material the sentence is written from and
(b) a detail sheet you read down instead of across.

Sheets:
    County digest     one row per county, every county, plain English
    Category detail   the 17 categories read vertically, with counties and examples
    Full matrix       the original wide grid, preserved so nothing is lost
    Site replaced     the election-night static swaps and 404s
    Numbers           figures that moved, with deltas
    Reading this      how to use it

Usage:
    python scripts/write_digest.py --from <rev> [--to <rev>]
"""

from __future__ import annotations

import argparse
import csv
from collections import Counter
from pathlib import Path

import analyze_diffs as A

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "manifest" / "fl-election-day-digest.xlsx"
FONT = "Arial"

# Category -> the verb phrase used when writing a county's sentence. Ordered
# roughly by how much a reader cares, which is NOT the same as how many lines each
# produces: a county posting results matters more than one relabelling its menu,
# even though the menu generates more lines.
PHRASE = {
    "Results posted": "posted election results",
    "Precincts reporting": "updated precinct reporting",
    "Turnout / ballots cast": "reported turnout",
    "Voter registration numbers": "updated registration figures",
    "Party / registration breakdown": "updated its party-registration breakdown",
    "Wait times / lines": "published polling-place wait times",
    "Election day / polling place": "updated election-day and polling-place info",
    "Early voting": "changed early-voting information",
    "Vote-by-mail / drop box": "updated vote-by-mail and drop-box info",
    "Sample ballot / candidates": "published sample ballots or candidate lists",
    "Provisional / cure": "added provisional-ballot and cure information",
    "Canvassing board": "posted canvassing-board or logic-and-accuracy notices",
    "Poll worker / staffing": "updated poll-worker information",
    "Next-election pivot": "started pointing ahead to the November general",
    "Alert / notice banner": "changed an alert or notice banner",
    "Date / countdown change": "moved dates or countdowns",
    "Navigation / menu labels": "relabelled its site navigation",
    "Translate widget (presentation)": "showed translate-widget churn",
}
# Substance ranking: which phrases lead a sentence. Everything not listed is
# treated as housekeeping and only mentioned if nothing better fired.
LEAD_PRIORITY = [
    "Results posted", "Precincts reporting", "Turnout / ballots cast",
    "Wait times / lines", "Voter registration numbers",
    "Party / registration breakdown", "Early voting",
    "Election day / polling place", "Sample ballot / candidates",
    "Vote-by-mail / drop box", "Provisional / cure", "Canvassing board",
    "Poll worker / staffing", "Next-election pivot", "Alert / notice banner",
    "Date / countdown change", "Navigation / menu labels",
    "Translate widget (presentation)",
]


def _join(items: list[str]) -> str:
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    return ", ".join(items[:-1]) + " and " + items[-1]


def sentence(county: str, stats: dict, cats: Counter, figures: list[dict]) -> str:
    """Write the plain-English description of what one county did."""
    n, cap = stats["files"], stats["captured"]
    added, removed = stats["added"], stats["removed"]

    fired = [c for c in LEAD_PRIORITY if cats.get(c)]
    # Name at most three, chosen by substance rank rather than line volume — the
    # biggest number is usually navigation, which is the least interesting thing.
    named = _join([PHRASE[c] for c in fired[:3]])

    shape = ""
    if added and removed:
        if added > removed * 3:
            shape = "mostly publishing new material"
        elif removed > added * 3:
            shape = "mostly retiring material"
    # "pages" agrees with the denominator, not the numerator: "1 of 5 pages changed".
    page_word = "page" if cap == 1 else "pages"
    lead = f"{n} of {cap} {page_word} changed"
    if shape:
        lead += f", {shape}"

    body = f"It {named}." if named else "No substantive category matched."

    return f"{lead}. {body}{_figure_phrase(figures)}"


def _figure_phrase(figures: list[dict]) -> str:
    """Render the most informative figure a county reported, if any.

    Prefers a value that actually moved, so the sentence carries a number rather
    than the label above it — "Total: 15,886 -> 15,906 (+20)" beats
    "Active Registered Voters as of 08/18/2026:", which is what naively taking the
    first pair gives you.
    """
    if not figures:
        return ""
    with_delta = [f for f in figures if f.get("delta") is not None]
    if with_delta:
        f = with_delta[0]
        sign = "+" if f["delta"] >= 0 else ""
        return (f' Figure moved: {f["before"].strip()[:44]} → '
                f'{f["after"].strip()[:44]} ({sign}{f["delta"]:,.0f}).')
    # No paired delta: fall back to the newly-reported line with the most digits.
    best = max(figures, key=lambda f: sum(c.isdigit() for c in f["after"]))
    if any(c.isdigit() for c in best["after"]):
        return f' Reported: "{best["after"].strip()[:110]}"'
    return ""


def build(rev_from: str, rev_to: str | None, out: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    a = A.analyze(rev_from, rev_to)
    captured, all_counties = A.load_manifest()

    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    base = Font(name=FONT, size=10)
    bold = Font(name=FONT, bold=True, size=10)
    border = Border(bottom=Side(style="thin", color="BFBFBF"))
    wrap = Alignment(wrap_text=True, vertical="top")

    status_fill = {
        "Edited": PatternFill("solid", fgColor="DDEBF7"),
        "Election-night page (old links 404)": PatternFill("solid", fgColor="FFC7CE"),
        "Election-night page (old links serve it)":
            PatternFill("solid", fgColor="FFEB9C"),
        "Not comparable": PatternFill("solid", fgColor="EDEDED"),
        "No change": PatternFill("solid", fgColor="FFFFFF"),
    }

    def sheet(ws, headers, widths, freeze="A2"):
        ws.append(headers)
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i)
            c.fill, c.font = head_fill, head_font
            c.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = freeze

    wb = Workbook()

    # ---------------------------------------------------------------- digest --
    ws = wb.active
    ws.title = "County digest"
    heads = ["County", "What happened", "Pages changed", "In plain English",
             "Categories that fired (with line counts)", "SOE site"]
    sheet(ws, heads, [15, 32, 12, 86, 58, 40])

    # Index the anomalies by county so every county gets exactly one row.
    lite = {}
    empt = {}
    flip = {}
    for o in a["outages"]:
        {"lite": lite, "empty": empt, "render_flip": flip}[o["class"]].setdefault(
            o["county"], []).append(o["page_type"])

    # County SOE homepage, for the clickable link column.
    home_url: dict[str, str] = {}
    with (ROOT / "manifest" / "targets.csv").open(newline="",
                                                  encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            if r["page_type"].strip() == "homepage":
                home_url[r["county"].strip()] = r["url"].strip()

    figures_by_county: dict[str, list[dict]] = {}
    for n in a["numbers"]:
        figures_by_county.setdefault(n["county"], []).append(n)

    rows = []
    for county in sorted(all_counties):
        cap = captured.get(county, 0)
        st = a["per_county_stats"].get(county)
        cats = a["per_county_cat"].get(county, Counter())
        figs = figures_by_county.get(county, [])

        # All ten of these counties did the SAME thing — swapped their site for a
        # static election-night page. They differ only in how the server treats the
        # old deep URLs, which is a side effect, not a second behaviour. Labelling
        # that difference as if it were two behaviours (an earlier version of this
        # script did) makes Calhoun and Dixie look unrelated when they are not.
        if county in empt:
            pages = ", ".join(pt.replace("_", " ") for pt in sorted(empt[county]))
            status = "Election-night page (old links 404)"
            desc = (f"Swapped its site for a minimal static election-night page "
                    f"linking straight to results, turnout and precinct lookup. "
                    f"Its {pages} {'URL' if len(empt[county]) == 1 else 'URLs'} "
                    f"now return 404, so a saved link breaks outright — which at "
                    f"least fails loudly.")
            changed = f"{len(empt[county])} now 404"
        elif county in lite:
            status = "Election-night page (old links serve it)"
            desc = ("Swapped its site for a minimal static election-night page "
                    "linking straight to results, turnout and precinct lookup. "
                    "Every old URL still returns HTTP 200 — but serves that same "
                    "election-night page instead of the page requested, so a saved "
                    "link silently gives you the wrong content.")
            changed = f"{len(lite[county])} replaced"
        elif county in flip:
            status = "Not comparable"
            desc = ("Nothing changed on the county's side. The two captures used "
                    "different fetch paths: the previous run was on GitHub Actions "
                    "(a datacenter IP), where a plain request to this site is "
                    "challenged and the pipeline escalates to a headless browser; "
                    "this run was local, where a plain request succeeds. A browser "
                    "render includes JS-built content a plain fetch never sees, so "
                    "the two are not comparable.")
            changed = f"{len(flip[county])} affected"
        elif st:
            status = "Edited"
            desc = sentence(county, {**st, "captured": cap}, cats, figs)
            changed = f"{st['files']} of {cap}"
        else:
            status = "No change"
            desc = "Nothing changed on any captured page."
            changed = f"0 of {cap}"

        catlist = " · ".join(f"{c} ({n})" for c, n in
                             sorted(cats.items(), key=lambda x: -x[1])) or "—"
        rows.append((county, status, changed, desc, catlist,
                     home_url.get(county, "")))

    order = ["Election-night page (old links 404)",
             "Election-night page (old links serve it)", "Edited",
             "Not comparable", "No change"]
    rows.sort(key=lambda r: (order.index(r[1]), r[0]))

    for r in rows:
        ws.append(list(r))
        i = ws.max_row
        if str(r[5]).startswith("http"):
            c = ws.cell(row=i, column=6)
            c.hyperlink = r[5]
            c.font = Font(name=FONT, size=10, color="0563C1", underline="single")
        for col in range(1, 6):
            c = ws.cell(row=i, column=col)
            c.font = base
            c.border = border
            c.alignment = wrap
        ws.cell(row=i, column=1).font = bold
        ws.cell(row=i, column=2).fill = status_fill.get(r[1], PatternFill())
        ws.row_dimensions[i].height = None
    ws.auto_filter.ref = f"A1:F{ws.max_row}"

    # ------------------------------------------------------- category detail --
    wc = wb.create_sheet("Category detail")
    heads2 = ["Category", "What it counts", "Lines", "Counties",
              "Which counties", "Example lines"]
    sheet(wc, heads2, [30, 44, 8, 10, 52, 74])
    WHAT = {
        "Results posted": "Returns being published or a results page changing.",
        "Precincts reporting": "How many precincts have reported.",
        "Turnout / ballots cast": "How many people have voted.",
        "Voter registration numbers": "Registration totals and 'as of' dates.",
        "Party / registration breakdown": "The party-split lines under a total.",
        "Election day / polling place": "Where to vote today; precinct finders.",
        "Early voting": "Early-voting dates, sites and hours.",
        "Vote-by-mail / drop box": "Mail ballots, drop boxes, intake stations.",
        "Sample ballot / candidates": "Sample ballots and candidate lists.",
        "Next-election pivot": "Content pointing ahead to the November general.",
        "Date / countdown change": "Any date or countdown moving.",
        "Alert / notice banner": "Alert and notice banners.",
        "Provisional / cure": "Provisional ballots and signature cures.",
        "Canvassing board": "Canvassing board, logic-and-accuracy, audits.",
        "Poll worker / staffing": "Poll-worker recruitment and information.",
        "Wait times / lines": "Queue length at polling places.",
        "Navigation / menu labels": "Site menus being re-rendered.",
        "Translate widget (presentation)":
            "A translate widget's language list. Should read 0 — anything here "
            "means presentation churn is leaking into the content counts.",
    }
    for label, _ in A.CATEGORY_PATTERNS:
        cos = sorted(a["cat_counties"].get(label, ()))
        exs = a["cat_examples"].get(label, [])[:3]
        wc.append([label, WHAT.get(label, ""), a["cat_total"].get(label, 0),
                   len(cos), ", ".join(cos) or "—",
                   "\n".join(f"· {e['line'][:110]}" for e in exs) or "—"])
        i = wc.max_row
        for col in range(1, 7):
            c = wc.cell(row=i, column=col)
            c.font = base
            c.border = border
            c.alignment = wrap
        wc.cell(row=i, column=1).font = bold

    # ------------------------------------------------------------ full matrix --
    wm = wb.create_sheet("Full matrix")
    labels = [c for c, _ in A.CATEGORY_PATTERNS]
    sheet(wm, ["County", "Pages changed", "Lines changed", "Lines added",
               "Lines removed"] + labels,
          [15, 12, 12, 11, 12] + [13] * len(labels))
    for county in sorted(a["per_county_stats"]):
        st = a["per_county_stats"][county]
        wm.append([county, st["files"], st["added"] + st["removed"], st["added"],
                   st["removed"]]
                  + [a["per_county_cat"][county].get(l, 0) for l in labels])
        i = wm.max_row
        for col in range(1, 6 + len(labels)):
            wm.cell(row=i, column=col).font = base
            wm.cell(row=i, column=col).border = border

    # ------------------------------------------------------- site replaced ----
    # Needs the manifest URL and the URL actually served, which live in the
    # manifest and in each capture's meta.json rather than in the diff.
    import json as _json
    manifest_url = {}
    with (ROOT / "manifest" / "targets.csv").open(newline="",
                                                  encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            manifest_url[(r["county"].strip(), r["page_type"].strip())] = \
                r["url"].strip()

    def _final_url(county: str, ptype: str) -> str:
        slug = county.lower().replace(" ", "_")
        p = ROOT / "snapshots" / slug / ptype / "meta.json"
        if not p.exists():
            return ""
        try:
            return _json.loads(p.read_text(encoding="utf-8")).get("final_url", "")
        except Exception:  # noqa: BLE001
            return ""

    wo = wb.create_sheet("Election-night switches")
    sheet(wo, ["County", "Page type", "What a saved link does now",
               "URL we asked for (old)", "URL actually served (new)",
               "HTTP", "Bytes before → after"],
          [14, 14, 40, 54, 54, 12, 18])
    explain = {
        "empty": "BREAKS — returns 404",
        "lite": "WRONG PAGE — 200, but serves the election-night page",
        "render_flip": "unchanged (captured differently, not comparable)",
    }
    for o in sorted(a["outages"], key=lambda x: (x["class"] != "empty",
                                                 x["county"], x["page_type"])):
        old_u = manifest_url.get((o["county"], o["page_type"]), "")
        new_u = _final_url(o["county"], o["page_type"])
        wo.append([o["county"], o["page_type"].replace("_", " "),
                   explain.get(o["class"], o["class"]), old_u, new_u,
                   o["status"], f'{o["bytes_before"]:,} → {o["bytes_after"]:,}'])
        i = wo.max_row
        for col in range(1, 8):
            c = wo.cell(row=i, column=col)
            c.font = base
            c.border = border
            c.alignment = wrap
        for col, url in ((4, old_u), (5, new_u)):
            if url.startswith("http"):
                cell = wo.cell(row=i, column=col)
                cell.hyperlink = url
                cell.font = Font(name=FONT, size=10, color="0563C1",
                                 underline="single")
        wo.cell(row=i, column=3).fill = (
            PatternFill("solid", fgColor="FFC7CE") if o["class"] == "empty"
            else PatternFill("solid", fgColor="FFEB9C") if o["class"] == "lite"
            else PatternFill("solid", fgColor="EDEDED"))

    # ------------------------------------------------------------- numbers ----
    wn = wb.create_sheet("Numbers")
    sheet(wn, ["County", "Page type", "Kind", "Before", "After", "Change"],
          [15, 14, 10, 60, 60, 11])
    for n in sorted(a["numbers"], key=lambda x: (x["kind"] != "changed",
                                                 x["county"])):
        wn.append([n["county"], n["page_type"], n["kind"], n["before"], n["after"],
                   n["delta"] if n.get("delta") is not None else ""])
        i = wn.max_row
        for col in range(1, 7):
            wn.cell(row=i, column=col).font = base
            wn.cell(row=i, column=col).border = border
            wn.cell(row=i, column=col).alignment = wrap
        if n.get("delta") is not None:
            wn.cell(row=i, column=6).number_format = "+#,##0;-#,##0;0"

    # -------------------------------------------------------- reading this ----
    wr = wb.create_sheet("Reading this")
    wr.column_dimensions["A"].width = 24
    wr.column_dimensions["B"].width = 108
    wr["A1"] = "How to read this workbook"
    wr["A1"].font = Font(name=FONT, bold=True, size=12)
    wr.append([])
    span_from = A._sh(["git", "log", "-1", "--format=%h  %ad  %s", "--date=short",
                       rev_from]).strip()
    guide = [
        ("Start here", "'County digest' — one row per county, all 67, sorted so the "
                       "counties that broke or replaced pages come first."),
        ("Comparison", f"{span_from}  →  {rev_to or 'fresh capture'}"),
        ("What happened", "Five states: 'Pages went 404' (old URLs stopped working) "
                          "· 'Switched to election-night page' · 'Edited' · 'Not "
                          "comparable' (captured a different way, not a real change) "
                          "· 'No change'."),
        ("In plain English", "A written description of what that county actually did. "
                             "Built from the categories, naming the most substantive "
                             "ones rather than the largest — a county posting results "
                             "matters more than one relabelling its menu, even though "
                             "the menu produces more lines."),
        ("Categories that fired",
         "Every category that matched, with its line count, biggest first. All 17 are "
         "kept; they are just listed in one cell instead of spread across 17 columns."),
        ("Category detail", "The same 17 read vertically: what each one counts, which "
                            "counties had it, and real example lines."),
        ("Full matrix", "The original wide grid, unchanged, so nothing is lost."),
        ("A caution on counts",
         "Category counts do NOT add up to lines changed, and are not meant to: a line "
         "matching two categories is counted twice, and a line matching none is not "
         "counted at all. Use them to compare counties on the SAME category, not to "
         "decompose a county's total."),
        ("Line counts vs counties",
         "A big line count can come from one verbose county. To judge whether "
         "something was widespread, read the 'Counties' column on 'Category detail'."),
        ("Regenerate", "python scripts/write_digest.py --from <rev>"),
    ]
    for k, v in guide:
        wr.append([k, v])
        i = wr.max_row
        wr.cell(row=i, column=1).font = bold
        wr.cell(row=i, column=2).font = base
        wr.cell(row=i, column=2).alignment = wrap

    wb.calculation.fullCalcOnLoad = True
    wb.save(out)

    # Also emit the digest as CSV, which reads fine in any tool.
    with out.with_suffix(".csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["county", "what_happened", "pages_changed", "plain_english",
                    "categories", "soe_site"])
        w.writerows(rows)

    tally = Counter(r[1] for r in rows)
    print(f"wrote {out}")
    print(f"wrote {out.with_suffix('.csv')}")
    print(f"\n{len(rows)} counties:")
    for k in order:
        if tally.get(k):
            print(f"  {k:<34} {tally[k]}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--from", dest="rev_from", default="HEAD")
    ap.add_argument("--to", dest="rev_to", default=None)
    ap.add_argument("--out", default=str(OUT))
    args = ap.parse_args()
    build(args.rev_from, args.rev_to, Path(args.out))


if __name__ == "__main__":
    main()
